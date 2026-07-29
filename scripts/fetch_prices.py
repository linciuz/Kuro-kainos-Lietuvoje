#!/usr/bin/env python3
"""
Kuro Kainos Lietuvoje - official price fetcher.

Source: Lietuvos energetikos agentūra (LEA), https://www.ena.lt/degalu-kainos-degalinese/
Stations must report 95 petrol / diesel / LPG prices to LEA every working day by 10:00.
LEA publishes a daily Excel with every station's prices (hosted on SharePoint).

This script:
  1. Reads the LEA page and finds the newest "Naujausios degalų kainos" Excel link.
  2. Downloads it anonymously via the OneDrive/SharePoint shares API.
  3. Parses it adaptively (matches Lithuanian header keywords; handles both
     "one column per fuel" and "one row per fuel" layouts).
  4. Writes data/stations.json (per-station rows + national summary).

It prints everything it detects (link, sheet, headers, column mapping, sample
rows) so the first GitHub Actions run shows exactly what came back. If a column
isn't mapped correctly, adjust the KEYWORDS below to match the real headers from
the log - no other code change needed.

NOTE: this has not been run against the live file from the dev environment
(that sandbox can't reach ena.lt/SharePoint). It is meant to run in GitHub
Actions, which has open internet. Treat the first run as a validation run.
"""

import base64
import datetime as dt
import json
import os
import re
import sys

import requests
from openpyxl import load_workbook

# Make stdout/stderr UTF-8 so Lithuanian text in the logs doesn't crash on a
# Windows console (cp1252). No-op where stdout is already UTF-8 (e.g. CI).
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

PAGE_URL = "https://www.ena.lt/degalu-kainos-degalinese/"
OUT_PATH = os.path.join("data", "stations.json")
DEBUG_HEADERS_PATH = os.path.join("data", "_debug_headers.json")
UA = "Mozilla/5.0 (compatible; KuroKainosBot/1.0; +https://github.com/)"
# A real browser UA for the PAGE fetch — some CDNs serve a cached/stale page to
# non-browser UAs, which can hide a file published minutes earlier.
BROWSER_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")

# Keyword -> field. Matching is case-insensitive and accent-insensitive on the
# Excel header text. Extend these lists if the log shows unmapped columns.
KEYWORDS = {
    "network":      ["tinkl", "imone", "prekes zenkl", "operatorius", "brand"],
    # NB: no bare "vieta" here - the LEA file has two "Degalines vieta (...)"
    # columns (Savivaldybe and Gyvenviete, gatve); "vieta" would greedily grab
    # the savivaldybe column and steal it from `municipality`.
    "address":      ["gatve", "gyvenviet", "adres", "degalines pavadinim"],
    "municipality": ["savivaldyb"],
    "locality":     ["miest", "kaim"],
    "fuel":         ["degalu tipas", "tipas", "degalu rus", "kuro rus", "produkt", "rusis"],  # long-format fuel column
    "price":        ["kaina"],                                        # long-format single price column
    "date":         ["pateikimo data", "data"],                       # date the price was reported
}

# Wide-format: a separate price column per fuel. Header keyword -> our fuel key.
FUEL_COLUMN_KEYWORDS = {
    "petrol95": ["95", "benzin", "e95", "a95"],
    "diesel":   ["dyzel", "disel", "d", "dt"],
    "lpg":      ["snd", "dujos", "lpg", "suskystint"],
}

# Long-format: map a fuel-cell value to our fuel key.
FUEL_VALUE_KEYWORDS = {
    "petrol95": ["95", "benzin"],
    "diesel":   ["dyzel", "diesel"],
    "lpg":      ["snd", "dujos", "lpg"],
}


def deaccent(s):
    repl = {"ą": "a", "č": "c", "ę": "e", "ė": "e", "į": "i", "š": "s",
            "ų": "u", "ū": "u", "ž": "z"}
    s = (s or "").lower()
    for a, b in repl.items():
        s = s.replace(a, b)
    return s


def station_key(net, addr, muni):
    """One station identity rule for BOTH price sources (Excel + portal), so the
    two can never disagree about what counts as the same physical station."""
    return f"{net}|{addr}|{muni}"


def blocks_price_candidates(html):
    """Candidates parsed from PARAGRAPH TEXT next to the anchors.

    2026-07-17 markup change: LEA's hand-edited CMS page moved the label
    ('Naujausios degalų kainos (YYYY-MM-DD)') OUT of the anchor — anchors are
    now empty, and their title= attributes carry stale copy-pasted dates. So:
    split into <p> blocks; a block whose plain text carries a dated 'naujausi…'
    label yields one candidate per SharePoint href in it (in order — the right
    one is usually first, but main() validates the downloaded file's own date
    against the label before accepting, so order only affects attempt count).
    """
    out = []
    for block in re.split(r"</p\s*>", html):
        hrefs = re.findall(r'href="(https://[^"]*sharepoint\.com/[^"]+)"', block)
        if not hrefs:
            continue
        text = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", block)).strip()
        td = deaccent(text)
        m = re.search(r"naujausi\w*[^()]{0,80}\((20\d\d-\d\d-\d\d)\)", td)
        if not m or "pranesim" in td:
            continue
        for href in hrefs:
            out.append({"href": href, "label": text[:90], "td": td, "date": m.group(1)})
    return out


def find_price_candidates(html):
    """Ordered candidate links for the CURRENT daily PRICE file (best first).
    Anchor-label parsing (old markup) first, then block-text parsing (new
    markup); within each, newest label date wins. main() download-validates."""
    anchors = re.findall(
        r'<a[^>]+href="(https://[^"]*sharepoint\.com/[^"]+)"[^>]*>(.*?)</a>',
        html, re.I | re.S)
    cands = []
    for href, text in anchors:
        label = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", text)).strip()
        m = re.search(r"(20\d\d-\d\d-\d\d)", label)
        cands.append({"href": href, "label": label,
                      "td": deaccent(label), "date": m.group(1) if m else ""})
    block_cands = blocks_price_candidates(html)

    print(f"[info] {len(cands)} SharePoint anchor(s), {len(block_cands)} block-text candidate(s):")
    for c in cands + block_cands:
        print(f"    - [{c['date'] or '          '}] {c['label'][:72]}")

    ordered, seen = [], set()

    def add(subset, why):
        for c in sorted(subset, key=lambda c: c["date"], reverse=True):
            if c["href"] not in seen:
                seen.add(c["href"])
                ordered.append(c)
                print(f"[info] candidate ({why}): [{c['date'] or '?'}] {c['href'][:70]}")

    # 1) anchor-labeled PRICE file (pre-2026-07-17 markup)
    add([c for c in cands if "naujausios" in c["td"] and "kainos" in c["td"]
         and "pranesim" not in c["td"]], "anchor label")
    # 2) block-text label beside the anchor (2026-07-17 markup)
    add(block_cands, "block text")
    # 3) any dated anchor except the 'pranešimas' report
    add([c for c in cands if c["date"] and "pranesim" not in c["td"]], "dated anchor")
    # NO blind first-link fallback: grabbing "whatever is first" published a
    # stale May snapshot once. Empty list -> caller fails safe, gates go loud.
    return ordered


def lea_price_dates(html):
    """All label dates of PRICE files visible on the page (both markups) —
    the ground truth verify_freshness compares our published date against."""
    dates = {c["date"] for c in blocks_price_candidates(html)}
    for href, text in re.findall(
            r'<a[^>]+href="(https://[^"]*sharepoint\.com/[^"]+)"[^>]*>(.*?)</a>', html, re.I | re.S):
        label = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", text)).strip()
        td = deaccent(label)
        m = re.search(r"(20\d\d-\d\d-\d\d)", label)
        if m and "naujausios" in td and "kain" in td and "pranesim" not in td:
            dates.add(m.group(1))
    return sorted(dates)


def find_latest_excel_link(html):
    """Return the SharePoint link for the CURRENT daily PRICE file.

    The page lists several dated files. Two look alike but are DIFFERENT:
      * 'Naujausios degalų kainos (YYYY-MM-DD)'        <- the daily PRICES  (use this)
      * 'Naujausias pranešimas apie degalų kainas (…)' <- an ANALYSIS report
    They usually share a date, but the report can LAG (e.g. after a holiday the
    prices update while the report stays a few days old). We must pick the
    latest-dated *prices* file and never the report — matching by label, and
    taking the newest date, not page position. Every candidate is logged so the
    run shows exactly what was on the page and which was chosen.
    """
    anchors = re.findall(
        r'<a[^>]+href="(https://[^"]*sharepoint\.com/[^"]+)"[^>]*>(.*?)</a>',
        html, re.I | re.S)

    cands = []
    for href, text in anchors:
        label = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", text)).strip()
        m = re.search(r"(20\d\d-\d\d-\d\d)", label)
        cands.append({"href": href, "label": label,
                      "td": deaccent(label), "date": m.group(1) if m else ""})

    print(f"[info] {len(cands)} SharePoint anchor(s) on the page:")
    for c in cands:
        print(f"    - [{c['date'] or '          '}] {c['label'][:72]}")

    def pick(subset, why):
        best = max(subset, key=lambda c: c["date"])   # newest date wins
        print(f"[info] selected {why}: [{best['date'] or '?'}] {best['label'][:60]}")
        return best["href"]

    # 1) The daily PRICES file: 'naujausios ... kainos', explicitly NOT the
    #    'pranešimas' report; take the newest date.
    price = [c for c in cands if "naujausios" in c["td"]
             and "kainos" in c["td"] and "pranesim" not in c["td"]]
    if price:
        return pick(price, "PRICE file (naujausios degalu kainos)")

    # 2) Any 'naujausios ... kain' anchor (excluding the report), newest date.
    loose = [c for c in cands if "naujausios" in c["td"]
             and "kain" in c["td"] and "pranesim" not in c["td"]]
    if loose:
        return pick(loose, "loose price match")

    # 3) Any dated anchor, newest date — still excluding the 'pranešimas' report,
    #    which is dated too and often newest (the exact 2026-07-07 failure mode).
    dated = [c for c in cands if c["date"] and "pranesim" not in c["td"]]
    if dated:
        return pick(dated, "newest dated anchor")

    # No blind fallback: the first SharePoint link on the page is a stale May
    # wide-format snapshot, so grabbing "whatever is first" would publish OLD
    # data with a green run. Better to fail (keeps last good file) and let the
    # freshness gates make the miss loud.
    print("[error] no dated 'naujausios ... kainos' anchor found — refusing to guess")
    return None


def _looks_like_xlsx(content):
    # .xlsx is a zip; it always starts with the PK signature.
    return content[:2] == b"PK"


def download_shared_xlsx(share_url):
    """Download an anonymously-shared SharePoint/OneDrive file.

    Primary method (works for SharePoint Online tenant share links like
    ltenergagen.sharepoint.com/:x:/...): append download=1 to the share URL,
    which makes SharePoint stream the raw file and follow redirects to it.
    Falls back to the consumer OneDrive shares API for personal-OneDrive links.
    """
    headers = {"User-Agent": UA}

    # Primary: download=1 on the share link itself (keeps the ?e= access token).
    sep = "&" if "?" in share_url else "?"
    direct = share_url + sep + "download=1"
    try:
        r = requests.get(direct, headers=headers, allow_redirects=True, timeout=60)
        if r.status_code == 200 and _looks_like_xlsx(r.content):
            return r.content
        print(f"[warn] download=1 returned status={r.status_code} "
              f"ct={r.headers.get('content-type','')[:40]} - trying fallback")
    except requests.RequestException as e:
        print(f"[warn] download=1 request failed: {e} - trying fallback")

    # Fallback: consumer OneDrive shares API (u! base64 of the share URL).
    enc = base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    api = f"https://api.onedrive.com/v1.0/shares/u!{enc}/driveItem/content"
    r = requests.get(api, headers=headers, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return r.content


def header_row_index(ws, max_scan=15):
    """Find the first row that looks like a header (>=3 non-empty text cells)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        nonempty = [c for c in row if c not in (None, "")]
        text = [c for c in nonempty if isinstance(c, str)]
        if len(nonempty) >= 3 and len(text) >= 3:
            return i
    return 0


def map_columns(headers):
    """Map column index -> our field name, using KEYWORDS.

    Each column is assigned to at most one role (first match wins) so a single
    column can't satisfy two fields - e.g. the savivaldybe column being read as
    both `address` and `municipality`, which would corrupt the per-station key
    and silently merge distinct stations.
    """
    mapping = {}
    fuel_cols = {}
    used = set()

    # Pass 1: wide-format fuel price columns (only if "kaina" or a fuel word present).
    for idx, h in enumerate(headers):
        hd = deaccent(str(h))
        if not hd.strip() or idx in used:
            continue
        for fuel, kws in FUEL_COLUMN_KEYWORDS.items():
            if fuel in fuel_cols:
                continue
            if any(k in hd for k in kws) and ("kaina" in hd or any(
                    f in hd for f in ["benzin", "dyzel", "dujos", "snd", "lpg"])):
                fuel_cols[fuel] = idx
                used.add(idx)
                break        # one column -> one fuel role only

    # Pass 2: generic text fields, each column claimed by at most one field.
    for idx, h in enumerate(headers):
        if idx in used:
            continue
        hd = deaccent(str(h))
        if not hd.strip():
            continue
        for field, kws in KEYWORDS.items():
            if field in mapping:
                continue
            if any(k in hd for k in kws):
                mapping[field] = idx
                used.add(idx)
                break
    return mapping, fuel_cols


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 3)
    s = str(v).strip().replace("€", "").replace("eur", "").replace("/l", "")
    s = s.replace(",", ".")
    m = re.search(r"\d+\.?\d*", s)
    return round(float(m.group()), 3) if m else None


def parse_workbook(xbytes):
    tmp = "_lea_tmp.xlsx"
    with open(tmp, "wb") as f:
        f.write(xbytes)
    wb = load_workbook(tmp, data_only=True, read_only=True)
    # pick the sheet with the most rows
    ws = max(wb.worksheets, key=lambda s: (s.max_row or 0))
    print(f"[info] sheets: {[s.title for s in wb.worksheets]}; using '{ws.title}' "
          f"({ws.max_row} rows x {ws.max_column} cols)")

    rows = list(ws.iter_rows(values_only=True))
    hidx = header_row_index(ws)
    headers = [str(c) if c is not None else "" for c in rows[hidx]]
    print(f"[info] header row #{hidx+1}: {headers}")

    mapping, fuel_cols = map_columns(headers)
    print(f"[info] field mapping: {mapping}")
    print(f"[info] wide fuel columns: {fuel_cols}")

    # save headers for debugging/validation
    os.makedirs("data", exist_ok=True)
    with open(DEBUG_HEADERS_PATH, "w", encoding="utf-8") as f:
        json.dump({"sheet": ws.title, "headers": headers,
                   "mapping": mapping, "fuel_cols": fuel_cols}, f,
                  ensure_ascii=False, indent=2)

    data_rows = rows[hidx + 1:]
    stations = {}

    key = station_key

    def cell(r, field):
        """Value of `field`'s column for row r, as a clean string ('' if blank).
        Blank Excel cells are None; without this they'd become the literal
        string 'None' and leak in as fake stations (e.g. footer rows)."""
        idx = mapping.get(field)
        if idx is None or idx >= len(r):
            return ""
        v = r[idx]
        return "" if v is None else str(v).strip()

    if fuel_cols:
        # WIDE format: one row per station, a price column per fuel
        for r in data_rows:
            if not r or all(c in (None, "") for c in r):
                continue
            net = cell(r, "network")
            addr = cell(r, "address")
            muni = cell(r, "municipality")
            loc = cell(r, "locality")
            # A real station always has a company; blank-company rows are the
            # spreadsheet's footer (national average / "Duomenys: N degalines").
            if not net:
                continue
            st = stations.setdefault(key(net, addr, muni), {
                "network": net, "address": addr, "municipality": muni,
                "locality": loc, "petrol95": None, "diesel": None, "lpg": None})
            for fuel, ci in fuel_cols.items():
                if ci < len(r):
                    val = to_float(r[ci])
                    if val:
                        st[fuel] = val
    else:
        # LONG format: one row per (station, fuel); needs fuel + price columns
        if "fuel" not in mapping or "price" not in mapping:
            print("[error] Could not identify fuel/price columns. "
                  "Check the header log above and update KEYWORDS.")
            sys.exit(2)
        for r in data_rows:
            if not r or all(c in (None, "") for c in r):
                continue
            net = cell(r, "network")
            addr = cell(r, "address")
            muni = cell(r, "municipality")
            loc = cell(r, "locality")
            fuel_raw = deaccent(cell(r, "fuel"))
            price = to_float(r[mapping["price"]]) if mapping["price"] < len(r) else None
            if not net or price is None:
                continue
            fuel = None
            for fk, kws in FUEL_VALUE_KEYWORDS.items():
                if any(k in fuel_raw for k in kws):
                    fuel = fk
                    break
            if not fuel:
                continue
            st = stations.setdefault(key(net, addr, muni), {
                "network": net, "address": addr, "municipality": muni,
                "locality": loc, "petrol95": None, "diesel": None, "lpg": None})
            st[fuel] = price

    # Detect the date the prices were actually reported, so we never again
    # label a stale file with today's date.
    file_date = None
    if "date" in mapping:
        ds = set()
        for r in data_rows:
            m = re.search(r"20\d\d-\d\d-\d\d", cell(r, "date"))
            if m:
                ds.add(m.group(0))
        if ds:
            file_date = max(ds)

    return list(stations.values()), file_date


# --- LEA's new self-service portal (primary source since 2026-07-28) ----------
# LEA restructured ena.lt on 2026-07-28: the daily SharePoint Excel link was
# removed from the page and the whole thing migrated to degalukainos.ena.lt, a
# Vue SPA backed by a public JSON API. Overnight the API's price coverage jumped
# 7% -> 94% and its prices matched the SharePoint file to the 3rd decimal, so it
# IS the daily feed now (plus official coords + per-station submit timestamps).
# We source prices from it FIRST and keep the SharePoint scraper as a fallback,
# because LEA still generates the Excel and the portal could yet wobble.
PORTAL_FUELS = {"benzinas_95": "petrol95", "dyzelinas": "diesel", "snd": "lpg"}


def portal_stations():
    """Return (stations, file_date) in parse_workbook's shape, sourced from the
    portal API. Raises on any failure or thin coverage so main() falls back to
    the SharePoint scraper. Prices only — coords come from the same downstream
    steps as before (geocode + chain snap + the lea_portal coord merge)."""
    import fetch_lea_portal as portal
    base, token = portal.discover_credentials()
    # portal.get() returns raw text, not parsed JSON.
    raw = json.loads(portal.get(f"{base}/read/prices?per_page=3000",
                                {"Authorization": f"Bearer {token}", "Accept": "application/json"}))
    rows = raw.get("data") or []
    if len(rows) < 500:
        raise RuntimeError(f"portal returned only {len(rows)} rows")

    stations, dates = {}, set()
    for r in rows:
        net = (r.get("company_name") or "").strip()
        addr = (r.get("address") or "").strip()
        muni = (r.get("municipality") or "").strip()
        if not net:
            continue
        st = stations.setdefault(station_key(net, addr, muni), {
            "network": net, "address": addr, "municipality": muni,
            "locality": "", "petrol95": None, "diesel": None, "lpg": None})
        fuel = PORTAL_FUELS.get(r.get("fuel_type"))
        price = to_float(r.get("price"))
        if fuel and price and 0.3 < price < 3.5:
            st[fuel] = price
        m = re.search(r"20\d\d-\d\d-\d\d", str(r.get("submitted_at") or ""))
        if m:
            dates.add(m.group(0))

    priced = [s for s in stations.values() if any(s[f] is not None for f in ("petrol95", "diesel", "lpg"))]
    if len(priced) < 400:
        raise RuntimeError(f"portal has only {len(priced)} priced stations (<400) — not usable yet")
    file_date = max(dates) if dates else None
    print(f"[ok] portal: {len(stations)} stations, {len(priced)} priced, newest submit {file_date}")
    return list(stations.values()), file_date


def _sharepoint_stations(cands):
    """FALLBACK source: the daily SharePoint Excel still linked from ena.lt.

    Try candidates in order; ACCEPT the first whose file's OWN date column
    matches its page label (LEA's anchors carry stale copy-paste titles and
    several links can share one paragraph — only the file itself is truth).
    Fall back to the newest-dated file actually downloaded, loudly.
    Exits the process if nothing usable downloads, so the last good data stays.
    """
    stations = file_date = None
    tried = []
    for c in cands[:5]:
        try:
            xbytes = download_shared_xlsx(c["href"])
            st, fd = parse_workbook(xbytes)
        except Exception as e:
            print(f"[warn] candidate download/parse failed ({type(e).__name__}: {e}) — next")
            continue
        if not st:
            print(f"[warn] candidate parsed 0 stations — next")
            continue
        tried.append((fd or "", st, c))
        if fd and c["date"] and fd == c["date"]:
            print(f"[ok] file date {fd} matches page label — accepted")
            return st, fd
        print(f"[warn] file date {fd!r} != label {c['date']!r} — trying next candidate")
    if tried:
        tried.sort(key=lambda t: t[0], reverse=True)
        fd, st, _c = tried[0]
        print(f"[warn] no candidate matched its label — using newest downloaded file ({fd or '?'}).")
        return st, (fd or None)
    print("[error] every candidate failed to download/parse - aborting, keeping last good data.")
    sys.exit(3)


def summarize(stations):
    out = {}
    for fuel in ("petrol95", "diesel", "lpg"):
        vals = [s[fuel] for s in stations if s.get(fuel)]
        if vals:
            out[fuel] = {"min": round(min(vals), 3),
                         "avg": round(sum(vals) / len(vals), 3),
                         "max": round(max(vals), 3),
                         "count": len(vals)}
    return out


def main():
    # PRIMARY: LEA's new portal API (see portal_stations). PRICE_SOURCE records
    # which fed the published file, so the freshness gate can tell them apart.
    stations = file_date = None
    source = "Lietuvos energetikos agentūra (ena.lt)"
    source_url = PAGE_URL
    # MULTI-SOURCE ENGINE (scripts/price_engine.py): poll every LEA channel and
    # keep the freshest price PER STATION AND PER FUEL. LEA's channels do not
    # update in lockstep — measured 2026-07-29, the portal carried 225 stations
    # fresher than that morning's spreadsheet (real intraday self-reports, some
    # at 12:39) while the spreadsheet led on 488. Racing them means a late or
    # broken channel can no longer hold the whole app back.
    # Diagnostics are committed (data/_price_source_debug.json) because the fetch
    # step is non-fatal and the Actions log API needs auth we do not have — that
    # combination once hid a 2-day outage behind green runs.
    import price_engine
    dbg = {"checked_utc": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat() + "Z"}
    stations = file_date = None
    engine_meta = None
    try:
        stations, file_date, engine_meta = price_engine.resolve()
        source_url = "https://degalukainos.ena.lt/"
        dbg.update(engine="ok", file_date=file_date,
                   stations=len(stations), winners=engine_meta.get("winner_counts"),
                   sources=engine_meta.get("sources"))
    except Exception as e:
        import traceback
        dbg.update(engine="FAILED", error=f"{type(e).__name__}: {e}",
                   trace=traceback.format_exc()[-700:])
        print(f"[error] every price source failed ({type(e).__name__}: {e}) — "
              f"keeping the last good data.")
    try:
        os.makedirs("data", exist_ok=True)
        json.dump(dbg, open(os.path.join("data", "_price_source_debug.json"), "w",
                            encoding="utf-8"), ensure_ascii=False, indent=1)
    except Exception:
        pass
    if stations is None:
        sys.exit(1)

    # NEVER REGRESS: measured 2026-07-17 — LEA's label said today but every
    # link in its paragraph served May archive snapshots. Whatever we picked,
    # refuse to replace committed data with an OLDER file (equal is fine: same
    # file re-fetched). The gates make persisting staleness loud; regressing
    # 2 months quietly is strictly worse than keeping yesterday.
    try:
        committed = json.load(open(OUT_PATH, encoding="utf-8")).get("updated") or ""
    except Exception:
        committed = ""
    if file_date and committed and file_date < committed:
        print(f"[error] newest fetchable file is {file_date} but committed data is {committed} — "
              f"refusing to regress. Keeping the newer committed data (LEA page mid-edit?).")
        sys.exit(5)
    print(f"[info] parsed {len(stations)} stations")
    for s in stations[:5]:
        print("   sample:", s)

    # Use the file's own date; warn loudly if it is stale (LEA sometimes leaves
    # old daily snapshots linked) or if averages look implausible.
    # NO today() fallback: stamping an undated workbook with today's date would
    # fake freshness and blind every downstream age gate. If the date column
    # vanishes (format change), fail and keep the last good file instead.
    if not file_date:
        print("[error] no date column detected — refusing to stamp today() over unknown-age data.")
        sys.exit(4)
    updated = file_date
    age = (dt.date.today() - dt.date.fromisoformat(file_date)).days
    print(f"[info] data date (from file): {file_date} ({age} day(s) old)")
    if age > 3:
        print(f"[warn] LEA file looks STALE ({age} days old) - check the source link!")

    summary = summarize(stations)
    p95 = summary.get("petrol95", {}).get("avg")
    if p95 and not (1.0 <= p95 <= 2.5):
        print(f"[warn] 95 average €{p95} is outside the plausible band - parser may be off.")

    payload = {
        "updated": updated,
        "source": source,
        "source_url": source_url,
        "price_engine": engine_meta,
        "summary": summary,
        "stations": sorted(stations, key=lambda s: (s.get("municipality") or "", s.get("network") or "")),
    }
    os.makedirs("data", exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[ok] wrote {OUT_PATH} ({len(stations)} stations)")


if __name__ == "__main__":
    main()
